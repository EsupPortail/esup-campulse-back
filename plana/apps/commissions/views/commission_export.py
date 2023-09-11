"""Views directly linked to commission exports."""
import csv
from tempfile import NamedTemporaryFile

from django.core.exceptions import ObjectDoesNotExist
from django.db import models
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import OpenApiParameter, extend_schema
from openpyxl import Workbook
from rest_framework import generics, response, status
from rest_framework.permissions import IsAuthenticated

from plana.apps.associations.models import Association
from plana.apps.commissions.models import Commission, CommissionFund, Fund
from plana.apps.institutions.models.institution import Institution
from plana.apps.projects.models import (
    Category,
    Project,
    ProjectCategory,
    ProjectCommissionFund,
)
from plana.apps.projects.serializers.project import ProjectSerializer
from plana.apps.users.models import User
from plana.utils import generate_pdf


class CommissionExport(generics.RetrieveAPIView):
    """/commissions/{id}/export route"""

    permission_classes = [IsAuthenticated]
    queryset = Project.visible_objects.all()
    serializer_class = ProjectSerializer

    @extend_schema(
        parameters=[
            OpenApiParameter(
                "mode",
                OpenApiTypes.STR,
                OpenApiParameter.QUERY,
                description="Export mode (xlsx, csv or pdf, csv by default).",
            ),
            OpenApiParameter(
                "project_ids",
                OpenApiTypes.STR,
                OpenApiParameter.QUERY,
                description="Filter by Projects IDs.",
            ),
        ],
        responses={
            status.HTTP_200_OK: ProjectSerializer,
            status.HTTP_401_UNAUTHORIZED: None,
            status.HTTP_404_NOT_FOUND: None,
        },
    )
    def get(self, request, *args, **kwargs):
        """Projects presented to the commission export."""
        mode = request.query_params.get("mode")
        project_ids = request.query_params.get("project_ids")

        queryset = self.get_queryset()
        commission_id = kwargs["pk"]

        try:
            commission = Commission.objects.get(id=kwargs["pk"])
        except ObjectDoesNotExist:
            return response.Response(
                {"error": _("Commission does not exist.")},
                status=status.HTTP_404_NOT_FOUND,
            )

        if not request.user.has_perm("projects.view_project_any_fund"):
            managed_funds = request.user.get_user_managed_funds()
            if managed_funds.count() > 0:
                user_funds_ids = managed_funds
            else:
                user_funds_ids = request.user.get_user_funds()
        else:
            user_funds_ids = Fund.objects.all().values_list("id")

        if not request.user.has_perm("projects.view_project_any_institution"):
            user_institutions_ids = request.user.get_user_managed_institutions()
        else:
            user_institutions_ids = Institution.objects.all().values_list("id")

        if not request.user.has_perm(
            "projects.view_project_any_fund"
        ) or not request.user.has_perm("projects.view_project_any_institution"):
            user_associations_ids = request.user.get_user_associations()
            user_projects_ids = Project.visible_objects.filter(
                models.Q(user_id=request.user.pk)
                | models.Q(association_id__in=user_associations_ids)
            ).values_list("id")

            queryset = queryset.filter(
                models.Q(id__in=user_projects_ids)
                | models.Q(
                    id__in=(
                        ProjectCommissionFund.objects.filter(
                            commission_fund_id__in=CommissionFund.objects.filter(
                                fund_id__in=user_funds_ids
                            ).values_list("id")
                        ).values_list("project_id")
                    )
                )
                | models.Q(
                    association_id__in=Association.objects.filter(
                        institution_id__in=user_institutions_ids
                    ).values_list("id")
                )
            )

        fields = [
            str(_("Project ID")),
            str(_("Project name")),
            str(_("Manual identifier")),
            str(_("Association name")),
            str(_("Student misc name")),
            str(_("Project start date")),
            str(_("Project end date")),
            str(_("First edition")),
            str(_("Categories")),
        ]

        funds = Fund.objects.all().order_by("acronym")
        for fund in funds:
            acronym = fund.acronym
            fields.append(str(_("Amount asked ") + acronym))
            fields.append(str(_("Amount earned ") + acronym))

        projects = queryset.filter(
            id__in=ProjectCommissionFund.objects.filter(
                commission_fund_id__in=CommissionFund.objects.filter(
                    commission_id=commission_id
                ).values("id")
            ).values("project_id")
        ).order_by("id")

        if project_ids is not None and project_ids != "":
            projects = projects.filter(id__in=project_ids.split(","))

        data = {"projects": []}
        http_response = None
        writer = None
        workbook = None
        worksheet = None
        filename = f"commission_{commission_id}_export"
        if mode is None or mode == "csv":
            http_response = HttpResponse(content_type="application/csv")
            http_response[
                "Content-Disposition"
            ] = f"Content-Disposition: attachment; filename={filename}.csv"
            writer = csv.writer(http_response, delimiter=";")
            writer.writerow([field for field in fields])
        elif mode == "xlsx":
            workbook = Workbook()
            worksheet = workbook.active
            for index_field, field in enumerate(fields):
                worksheet.cell(row=1, column=(index_field + 1)).value = field
        elif mode == "pdf":
            data["name"] = commission.name
            data["fields"] = fields

        for index_project, project in enumerate(projects):
            association = (
                None
                if project.association_id is None
                else Association.objects.get(id=project.association_id).name
            )

            if project.user_id is None:
                user = None
            else:
                user = User.objects.get(id=project.user_id)
                user = f"{user.last_name} {user.first_name}"

            categories = list(
                Category.objects.filter(
                    id__in=ProjectCategory.objects.filter(
                        project_id=project.id
                    ).values_list("category_id")
                ).values_list("name", flat=True)
            )
            categories = ', '.join(categories)

            project_commission_funds = ProjectCommissionFund.objects.filter(
                project_id=project.id
            )

            is_first_edition = str(_("Yes"))
            for edition in project_commission_funds:
                if not edition.is_first_edition:
                    is_first_edition = str(_("No"))
                    break

            fields = [
                project.id,
                project.name,
                project.manual_identifier,
                association,
                user,
                project.planned_start_date.date(),
                project.planned_end_date.date(),
                is_first_edition,
                categories,
            ]
            for fund in funds:
                try:
                    pcf = ProjectCommissionFund.objects.get(
                        project_id=project.id,
                        commission_fund_id=CommissionFund.objects.get(
                            commission_id=commission_id, fund_id=fund.id
                        ).id,
                    )
                    fields.append(pcf.amount_asked)
                    fields.append(pcf.amount_earned)
                except ObjectDoesNotExist:
                    fields.append(0)
                    fields.append(0)

            if mode is None or mode == "csv":
                # Write CSV file content
                writer.writerow([field for field in fields])
            elif mode == "xlsx":
                for index_field, field in enumerate(fields):
                    worksheet.cell(
                        row=(index_project + 2), column=(index_field + 1)
                    ).value = field
            elif mode == "pdf":
                data["projects"].append(fields)

        if mode is None or mode == "csv":
            return http_response
        elif mode == "xlsx":
            with NamedTemporaryFile() as tmp:
                workbook.save(tmp.name)
                tmp.seek(0)
                stream = tmp.read()
            http_response = HttpResponse(
                content=stream,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            http_response[
                "Content-Disposition"
            ] = f"Content-Disposition: attachment; filename={filename}.xlsx"
            return http_response
        elif mode == "pdf":
            return generate_pdf(
                data["name"],
                data,
                "commission_projects_list",
                request.build_absolute_uri("/"),
            )
