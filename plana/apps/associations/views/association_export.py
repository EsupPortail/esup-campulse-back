"""Views directly linked to association exports."""
import csv
from tempfile import NamedTemporaryFile

from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Q
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import OpenApiParameter, extend_schema
from openpyxl import Workbook
from rest_framework import generics, response, status
from rest_framework.permissions import DjangoModelPermissions, IsAuthenticated

from plana.apps.associations.models.activity_field import ActivityField
from plana.apps.associations.models.association import Association
from plana.apps.associations.serializers.association import (
    AssociationAllDataReadSerializer,
)
from plana.apps.documents.models.document import Document
from plana.apps.documents.models.document_upload import DocumentUpload
from plana.apps.institutions.models import Institution, InstitutionComponent
from plana.apps.users.models import GroupInstitutionFundUser
from plana.utils import generate_pdf


class AssociationListExport(generics.RetrieveAPIView):
    """/associations/export route"""

    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    queryset = Association.objects.all()
    serializer_class = AssociationAllDataReadSerializer

    @extend_schema(
        operation_id="associations_export_list",
        parameters=[
            OpenApiParameter(
                "mode",
                OpenApiTypes.STR,
                OpenApiParameter.QUERY,
                description="Export mode (xlsx or csv, csv by default).",
            ),
            OpenApiParameter(
                "associations",
                OpenApiTypes.STR,
                OpenApiParameter.QUERY,
                description="IDs of selected associations, separated by a comma.",
            ),
        ],
        responses={
            status.HTTP_200_OK: AssociationAllDataReadSerializer,
            status.HTTP_401_UNAUTHORIZED: None,
            status.HTTP_403_FORBIDDEN: None,
        },
    )
    def get(self, request, *args, **kwargs):
        """Associations list export."""
        mode = request.query_params.get("mode")
        associations = request.query_params.get("associations")

        if request.user.has_perm("associations.change_association_any_institution"):
            queryset = self.get_queryset()
        else:
            institutions = GroupInstitutionFundUser.objects.filter(
                user_id=request.user.id, institution_id__isnull=False
            ).values_list("institution_id")
            if len(institutions) == 0:
                return response.Response(
                    {"error": _("Not allowed to export associations list CSV.")},
                    status=status.HTTP_403_FORBIDDEN,
                )
            queryset = self.get_queryset().exclude(~Q(institution_id__in=institutions))

        if associations is not None and associations != "":
            association_ids = [int(association) for association in associations.split(",")]
            queryset = queryset.exclude(~Q(id__in=association_ids))

        fields = [
            str(_("Name")),
            str(_("Acronym")),
            str(_("Institution")),
            str(_("Activity field")),
            str(_("Institution component")),
            str(_("Last GOA date")),
            str(_("Email")),
        ]

        http_response = None
        writer = None
        workbook = None
        worksheet = None
        filename = "associations_export"

        if mode is None or mode == "csv":
            http_response = HttpResponse(content_type="application/csv")
            http_response["Content-Disposition"] = f"Content-Disposition: attachment; filename={filename}.csv"
            writer = csv.writer(http_response, delimiter=";")
            writer.writerow([field for field in fields])
        elif mode == "xlsx":
            workbook = Workbook()
            worksheet = workbook.active
            for index_field, field in enumerate(fields):
                worksheet.cell(row=1, column=(index_field + 1)).value = field

        # Write CSV file content
        for index_association, association in enumerate(queryset):
            institution_component = (
                None
                if association.institution_component_id is None
                else InstitutionComponent.objects.get(id=association.institution_component_id).name
            )

            fields = [
                association.name,
                association.acronym,
                Institution.objects.get(id=association.institution_id).name,
                str(association.activity_field),
                institution_component,
                association.last_goa_date,
                association.email,
            ]

            if mode is None or mode == "csv":
                # Write CSV file content
                writer.writerow([field for field in fields])
            elif mode == "xlsx":
                for index_field, field in enumerate(fields):
                    worksheet.cell(row=(index_association + 2), column=(index_field + 1)).value = field

        if mode is None or mode == "csv":
            return http_response
        elif mode == "xlsx":
            with NamedTemporaryFile() as tmp:
                workbook.save(tmp.name)
                tmp.seek(0)
                stream = tmp.read()
            http_response = HttpResponse(
                content=stream,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            http_response["Content-Disposition"] = f"Content-Disposition: attachment; filename={filename}.xlsx"
            return http_response


class AssociationRetrieveExport(generics.RetrieveAPIView):
    """/associations/{id}/export route"""

    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    queryset = Association.objects.all()
    serializer_class = AssociationAllDataReadSerializer

    @extend_schema(
        operation_id="associations_export_retrieve",
        responses={
            status.HTTP_200_OK: AssociationAllDataReadSerializer,
            status.HTTP_401_UNAUTHORIZED: None,
            status.HTTP_403_FORBIDDEN: None,
            status.HTTP_404_NOT_FOUND: None,
        },
    )
    def get(self, request, *args, **kwargs):
        """Retrieves a PDF file."""
        try:
            association = self.queryset.get(id=kwargs["pk"])
            data = association.__dict__
        except ObjectDoesNotExist:
            return response.Response(
                {"error": _("Association does not exist.")},
                status=status.HTTP_404_NOT_FOUND,
            )

        if (
            not request.user.has_perm("associations.view_association_not_enabled")
            and not request.user.has_perm("associations.view_association_not_public")
            and not request.user.is_president_in_association(association.id)
        ):
            return response.Response(
                {"error": _("Not allowed to retrieve this association.")},
                status=status.HTTP_403_FORBIDDEN,
            )

        data["institution"] = Institution.objects.get(id=association.institution_id).name
        data["institution_component"] = InstitutionComponent.objects.get(id=association.institution_component_id).name
        data["activity_field"] = ActivityField.objects.get(id=association.activity_field_id).name

        data["documents"] = list(
            DocumentUpload.objects.filter(
                association_id=data["id"],
                document_id__in=Document.objects.filter(
                    process_type__in=["CHARTER_ASSOCIATION", "DOCUMENT_ASSOCIATION"]
                ),
            ).values("name", "document__name")
        )

        return generate_pdf(
            data["name"],
            data,
            "association_charter_summary",
            request.build_absolute_uri("/"),
        )
