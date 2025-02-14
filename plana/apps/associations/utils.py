import csv
from tempfile import NamedTemporaryFile

from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook

from plana.apps.institutions.models import Institution, InstitutionComponent


def generate_associations_export(queryset, mode: str) -> HttpResponse:
    fields = [
        str(_("Name")),
        str(_("Acronym")),
        str(_("Institution")),
        str(_("Activity field")),
        str(_("Institution component")),
        str(_("Charter date")),
        str(_("Last GOA date")),
        str(_("Email")),
    ]

    http_response = None
    writer = None
    workbook = None
    worksheet = None
    filename = "associations_export"

    if mode is None or mode == "csv":
        http_response = HttpResponse(content_type="application/csv")
        http_response["Content-Disposition"] = f"Content-Disposition: attachment; filename={filename}.csv"
        writer = csv.writer(http_response, delimiter=";")
        writer.writerow([field for field in fields])
    elif mode == "xlsx":
        workbook = Workbook()
        worksheet = workbook.active
        for index_field, field in enumerate(fields):
            worksheet.cell(row=1, column=index_field + 1).value = field

    # Write CSV file content
    for index_association, association in enumerate(queryset):
        institution_component = (
            None
            if association.institution_component_id is None
            else InstitutionComponent.objects.get(id=association.institution_component_id).name
        )

        fields = [
            association.name,
            association.acronym,
            Institution.objects.get(id=association.institution_id).name,
            str(association.activity_field),
            institution_component,
            association.charter_date,
            association.last_goa_date,
            association.email,
        ]

        if mode is None or mode == "csv":
            # Write CSV file content
            writer.writerow([field for field in fields])
        elif mode == "xlsx":
            for index_field, field in enumerate(fields):
                worksheet.cell(row=index_association + 2, column=index_field + 1).value = field

    if mode is None or mode == "csv":
        return http_response
    if mode == "xlsx":
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        http_response = HttpResponse(
            content=stream,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        http_response["Content-Disposition"] = f"Content-Disposition: attachment; filename={filename}.xlsx"
        return http_response
